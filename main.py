import os
import json
import time
import asyncio
import logging
import sqlite3
import re
from contextlib import closing
from typing import List, Set, Optional

try:
    from openpyxl import load_workbook
except Exception:
    load_workbook = None

from aiogram import Bot, Dispatcher, F
from aiogram.client.default import DefaultBotProperties
from aiogram.enums import ParseMode
from aiogram.filters import Command, CommandStart
from aiogram.types import (
    Message,
    WebAppInfo,
    MenuButtonWebApp,
    ReplyKeyboardMarkup,
    KeyboardButton,
    Document,
)

logging.basicConfig(level=logging.INFO, format="%(levelname)s:%(name)s:%(message)s")
logger = logging.getLogger(__name__)

print("=== LUX ARISTOKRAT QR BOT FIXED VERSION ===")

BOT_TOKEN = (
    os.getenv("BOT_TOKEN")
    or os.getenv("API_TOKEN")
    or os.getenv("TELEGRAM_BOT_TOKEN")
    or os.getenv("BOT_API_TOKEN")
    or ""
).strip()

WEBAPP_URL = (
    os.getenv("WEBAPP_URL", "").strip()
    or "https://tahirovdd-lang.github.io/QR-project-Lux-Aristokrat/?v=1"
)

DB_PATH = os.getenv("DB_PATH", "lux_aristokrat.db").strip() or "lux_aristokrat.db"

ADMIN_ID_RAW = os.getenv("ADMIN_ID", "0").strip()
ADMIN_IDS_RAW = os.getenv("ADMIN_IDS", "").strip()

QR_CODES_DIR = os.getenv("QR_CODES_DIR", "data/qr_codes").strip() or "data/qr_codes"
QR_CODES_DEFAULT_FILE = os.path.join(QR_CODES_DIR, "default_qr_codes.txt")

DEFAULT_ADMIN_IDS = {
    6013591658,
    6292063248,
}

if not BOT_TOKEN:
    raise RuntimeError("BOT_TOKEN / API_TOKEN / TELEGRAM_BOT_TOKEN / BOT_API_TOKEN is not set")


def parse_admin_ids() -> List[int]:
    result: Set[int] = set(DEFAULT_ADMIN_IDS)

    try:
        admin_id = int(ADMIN_ID_RAW)
        if admin_id > 0:
            result.add(admin_id)
    except Exception:
        pass

    if ADMIN_IDS_RAW:
        for part in ADMIN_IDS_RAW.split(","):
            part = part.strip()
            if not part:
                continue
            try:
                value = int(part)
                if value > 0:
                    result.add(value)
            except Exception:
                continue

    return sorted(result)


ADMIN_IDS = parse_admin_ids()

logging.info("WEBAPP_URL = %s", WEBAPP_URL)
logging.info("DB_PATH = %s", DB_PATH)
logging.info("QR_CODES_DIR = %s", QR_CODES_DIR)
logging.info("ADMIN_ID from env: %s", ADMIN_ID_RAW)
logging.info("ADMIN_IDS effective: %s", ADMIN_IDS)

# ВАЖНО: без ручного AiohttpSession/TCPConnector.
# Это исправляет RuntimeError: no running event loop.
bot = Bot(
    token=BOT_TOKEN,
    default=DefaultBotProperties(parse_mode=ParseMode.HTML),
)

dp = Dispatcher()

RECENT_SCANS = {}
SCAN_TTL_SECONDS = 8


TEXTS = {
    "ru": {
        "welcome": "Добро пожаловать в <b>Lux Aristokrat</b>\n\nОткройте mini app для сканирования QR / Data Matrix.",
        "open_scanner": "Открыть сканер",
        "menu_button": "Сканер QR",
        "empty_code": "Пустой код.",
        "bad_data": "Ошибка чтения данных.",
        "processing_error": "Ошибка обработки QR-кода.",
        "unknown_action": "Неизвестное действие.",
        "admin_only": "Команда только для администратора.",
        "debug_url": "Текущий WEBAPP_URL:\n{url}",
        "help": "Доступные команды:\n/start — старт\n/help — помощь\n/debug_url — показать WebApp URL\n/id — показать Telegram ID\n/addqr CODE 10 — добавить QR\n/delqr CODE — удалить QR\n/listqr — список QR\n/syncqr — обновить QR из папки data/qr_codes\n/importqr — инструкция импорта TXT/XLSX\n/bonus USER_ID 50 — начислить бонусы\n/balance USER_ID — баланс пользователя\n/check @username — проверить баллы и статус клиента по никнейму\n/status @username — проверить статус клиента по никнейму",
        "your_id": "Ваш Telegram ID: <code>{user_id}</code>\nAdmin access: <b>{admin}</b>",
        "cmd_usage_addqr": "Использование: <code>/addqr CODE 10</code>",
        "cmd_usage_delqr": "Использование: <code>/delqr CODE</code>",
        "cmd_usage_bonus": "Использование: <code>/bonus USER_ID 50</code>",
        "cmd_usage_balance": "Использование: <code>/balance USER_ID</code>",
        "cmd_usage_check": "Использование: <code>/check @username</code> или <code>/status @username</code>",
        "client_not_found": "Клиент <b>@{username}</b> не найден в базе. Клиент должен хотя бы один раз открыть бота через /start или сканер.",
        "client_info": "👤 Клиент: <b>{full_name}</b>\nUsername: <b>@{username}</b>\nTelegram ID: <code>{user_id}</code>\nБаллы: <b>{balance}</b>\nСтатус: <b>{status}</b>",
        "client_status": "👤 Клиент: <b>{full_name}</b>\nUsername: <b>@{username}</b>\nСтатус: <b>{status}</b>\nБаллы: <b>{balance}</b>",
        "addqr_ok": "QR добавлен:\n<code>{code}</code>\nБонусов: <b>{points}</b>",
        "addqr_exists": "QR уже существует:\n<code>{code}</code>",
        "delqr_ok": "QR удалён:\n<code>{code}</code>",
        "delqr_not_found": "QR не найден:\n<code>{code}</code>",
        "listqr_empty": "Список QR пуст.",
        "listqr_header": "Список QR-кодов:\n\n{items}",
        "bonus_ok": "Начислено <b>{amount}</b> бонусов пользователю <code>{user_id}</code>.\nНовый баланс: <b>{balance}</b>",
        "balance_text": "Баланс пользователя <code>{user_id}</code>: <b>{balance}</b>",
        "invalid_number": "Некорректное число.",
        "internal_error": "Внутренняя ошибка.",
    },
    "uz": {},
    "en": {},
    "tj": {},
}

TEXTS["uz"] = TEXTS["ru"].copy()
TEXTS["en"] = TEXTS["ru"].copy()
TEXTS["tj"] = TEXTS["ru"].copy()

TEXTS["uz"].update({
    "welcome": "<b>Lux Aristokrat</b> ga xush kelibsiz\n\nQR / Data Matrix skanerlash uchun mini app ni oching.",
    "open_scanner": "Skanerni ochish",
    "menu_button": "QR skaner",
})

TEXTS["en"].update({
    "welcome": "Welcome to <b>Lux Aristokrat</b>\n\nOpen the mini app to scan QR / Data Matrix.",
    "open_scanner": "Open scanner",
    "menu_button": "QR Scanner",
})

TEXTS["tj"].update({
    "welcome": "Хуш омадед ба <b>Lux Aristokrat</b>\n\nБарои скан кардани QR / Data Matrix mini app-ро кушоед.",
    "open_scanner": "Кушодани сканер",
    "menu_button": "Сканери QR",
})


SCAN_TEXTS = {
    "ru": {
        "scan_ok_title": "✅ Бонусы начислены",
        "scan_ok_text": "{title}\n\nКод: <code>{code}</code>\nНачислено бонусов: <b>{points}</b>\nВаш баланс: <b>{balance}</b>",
        "scan_used_title": "⚠️ Код уже использован",
        "scan_used_text": "{title}\n\nКод: <code>{code}</code>\nЭтот QR-код уже был активирован ранее.",
        "scan_invalid_title": "❌ Неверный код",
        "scan_invalid_text": "{title}\n\nКод: <code>{code}</code>\nТакой QR-код не найден или недоступен.",
    }
}

SCAN_TEXTS["uz"] = SCAN_TEXTS["ru"].copy()
SCAN_TEXTS["en"] = SCAN_TEXTS["ru"].copy()
SCAN_TEXTS["tj"] = SCAN_TEXTS["ru"].copy()


def normalize_lang(lang_code: str) -> str:
    lang_code = (lang_code or "").strip().lower()

    if lang_code.startswith("ru"):
        return "ru"
    if lang_code.startswith("uz"):
        return "uz"
    if lang_code.startswith("en"):
        return "en"
    if lang_code.startswith("tg") or lang_code.startswith("tj"):
        return "tj"

    return "ru"


def get_user_lang(message: Message) -> str:
    try:
        return normalize_lang((message.from_user.language_code if message.from_user else "ru") or "ru")
    except Exception:
        return "ru"


def t(lang: str, key: str, **kwargs) -> str:
    text = (TEXTS.get(lang) or TEXTS["ru"]).get(key) or TEXTS["ru"].get(key) or key
    return text.format(**kwargs) if kwargs else text


def scan_t(lang: str, key: str, **kwargs) -> str:
    text = (SCAN_TEXTS.get(lang) or SCAN_TEXTS["ru"]).get(key) or SCAN_TEXTS["ru"].get(key) or key
    return text.format(**kwargs) if kwargs else text


def is_admin(user_id: int) -> bool:
    return user_id in ADMIN_IDS


def normalize_username(username: str) -> str:
    return (username or "").strip().lstrip("@").lower()


def make_full_name(message: Message) -> str:
    user = message.from_user
    if not user:
        return ""

    parts = [
        (user.first_name or "").strip(),
        (user.last_name or "").strip(),
    ]

    return " ".join([p for p in parts if p]).strip()


def get_client_status(balance: int) -> str:
    balance = int(balance or 0)

    if balance >= 100:
        return "Gold"
    if balance >= 50:
        return "Silver"

    return "Bronze"


def update_user_profile(message: Message, language: str = "ru") -> None:
    user = message.from_user
    if not user:
        return

    username = normalize_username(user.username or "")
    full_name = make_full_name(message)

    with closing(get_db_connection()) as conn:
        ensure_loyalty_user(
            conn=conn,
            user_id=int(user.id),
            language=language,
            username=username,
            full_name=full_name,
        )
        conn.commit()


def build_main_keyboard(lang: str) -> ReplyKeyboardMarkup:
    return ReplyKeyboardMarkup(
        keyboard=[
            [
                KeyboardButton(
                    text=t(lang, "open_scanner"),
                    web_app=WebAppInfo(url=WEBAPP_URL),
                )
            ]
        ],
        resize_keyboard=True,
    )


def cleanup_recent_scans():
    now = time.time()
    for k in [k for k, ts in RECENT_SCANS.items() if now - ts > SCAN_TTL_SECONDS]:
        RECENT_SCANS.pop(k, None)


def is_duplicate_scan(user_id: int, code: str) -> bool:
    cleanup_recent_scans()

    key = f"{user_id}:{code}"
    now = time.time()

    if key in RECENT_SCANS:
        return True

    RECENT_SCANS[key] = now
    return False


def ensure_qr_codes_folder():
    os.makedirs(QR_CODES_DIR, exist_ok=True)

    if not os.path.exists(QR_CODES_DEFAULT_FILE):
        with open(QR_CODES_DEFAULT_FILE, "w", encoding="utf-8") as f:
            f.write("# Формат строки: КОД 5 баллов\n")
            f.write("# Пример: 047000280627027ABC 5 баллов\n")


def load_qr_codes_raw_from_folder() -> str:
    ensure_qr_codes_folder()

    parts = []

    try:
        file_names = sorted(os.listdir(QR_CODES_DIR))
    except Exception as e:
        logging.exception("Cannot read QR_CODES_DIR=%s: %s", QR_CODES_DIR, e)
        return ""

    for file_name in file_names:
        if not file_name.lower().endswith(".txt"):
            continue

        file_path = os.path.join(QR_CODES_DIR, file_name)

        if not os.path.isfile(file_path):
            continue

        try:
            with open(file_path, "r", encoding="utf-8") as f:
                parts.append(f"\n# ===== FILE: {file_name} =====\n")
                parts.append(f.read())
        except Exception as e:
            logging.exception("Cannot read QR file %s: %s", file_path, e)

    return "\n".join(parts)


def parse_default_qr_codes(raw: str):
    codes = []
    seen = set()

    for line in raw.splitlines():
        line = line.strip()

        if not line or line.startswith("#"):
            continue

        m = re.match(r"^(.*?)\s+(\d+)\s+бал", line, flags=re.IGNORECASE)

        if not m:
            m = re.match(r"^(.*?)\s+(\d+)$", line, flags=re.IGNORECASE)

        if not m:
            logging.warning("Skipped QR line: %r", line)
            continue

        code = m.group(1).strip()
        points = int(m.group(2))

        if code and code not in seen:
            seen.add(code)
            codes.append((code, points))

    return codes


def parse_qr_txt_content(content: str):
    return parse_default_qr_codes(content)


def parse_qr_xlsx_file(file_path: str):
    if load_workbook is None:
        raise RuntimeError("openpyxl is not installed. Add openpyxl==3.1.5 to requirements.txt or upload TXT file.")

    codes = []
    seen = set()

    wb = load_workbook(file_path, read_only=True, data_only=True)

    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            values = [v for v in row if v is not None]

            if not values:
                continue

            code = None
            points = None

            if len(values) >= 2:
                code = str(values[0]).strip()
                raw_points = str(values[1]).strip()
                m = re.search(r"\d+", raw_points)
                if m:
                    points = int(m.group(0))

            if not code or points is None:
                line = " ".join(str(v).strip() for v in values)
                m = re.match(r"^(.*?)\s+(\d+)\s+бал", line, flags=re.IGNORECASE)
                if m:
                    code = m.group(1).strip()
                    points = int(m.group(2))

            if not code or points is None:
                continue

            if code.lower() in {"код", "qr", "qr code", "data matrix", "балл", "баллы", "points"}:
                continue

            if code and code not in seen and points > 0:
                seen.add(code)
                codes.append((code, points))

    wb.close()
    return codes


def safe_file_name(name: str) -> str:
    name = os.path.basename(name or "qr_upload")
    name = re.sub(r"[^a-zA-Z0-9а-яА-ЯёЁ._-]+", "_", name)
    return name.strip("._") or "qr_upload"


def get_db_connection():
    conn = sqlite3.connect(DB_PATH, timeout=30, check_same_thread=False)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL;")
    conn.execute("PRAGMA synchronous=NORMAL;")
    conn.execute("PRAGMA foreign_keys=ON;")
    return conn


def init_qr_bonus_tables():
    with closing(get_db_connection()) as conn:
        cur = conn.cursor()

        cur.execute("""
            CREATE TABLE IF NOT EXISTS loyalty_users (
                user_id INTEGER PRIMARY KEY,
                username TEXT,
                full_name TEXT,
                language TEXT DEFAULT 'ru',
                bonus_balance INTEGER NOT NULL DEFAULT 0,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)

        try:
            cur.execute("ALTER TABLE loyalty_users ADD COLUMN username TEXT")
        except sqlite3.OperationalError:
            pass

        try:
            cur.execute("ALTER TABLE loyalty_users ADD COLUMN full_name TEXT")
        except sqlite3.OperationalError:
            pass

        cur.execute("""
            CREATE TABLE IF NOT EXISTS qr_codes (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                code TEXT NOT NULL UNIQUE,
                points INTEGER NOT NULL DEFAULT 1,
                is_active INTEGER NOT NULL DEFAULT 1,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)

        cur.execute("""
            CREATE TABLE IF NOT EXISTS qr_code_redemptions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                code TEXT NOT NULL,
                user_id INTEGER NOT NULL,
                points INTEGER NOT NULL DEFAULT 0,
                redeemed_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(code)
            )
        """)

        cur.execute("""
            CREATE TABLE IF NOT EXISTS bonus_transactions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                amount INTEGER NOT NULL,
                type TEXT NOT NULL,
                code TEXT,
                comment TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)

        cur.execute("CREATE INDEX IF NOT EXISTS idx_loyalty_users_username ON loyalty_users(username)")
        cur.execute("CREATE INDEX IF NOT EXISTS idx_qr_codes_code ON qr_codes(code)")
        cur.execute("CREATE INDEX IF NOT EXISTS idx_redemptions_user_id ON qr_code_redemptions(user_id)")
        cur.execute("CREATE INDEX IF NOT EXISTS idx_bonus_transactions_user_id ON bonus_transactions(user_id)")

        conn.commit()


def upsert_qr_codes_to_db(codes: list[tuple[str, int]]) -> dict:
    added = 0
    updated = 0
    skipped = 0

    with closing(get_db_connection()) as conn:
        cur = conn.cursor()

        for code, points in codes:
            code = str(code).strip()
            points = int(points)

            if not code or points <= 0:
                skipped += 1
                continue

            old = cur.execute(
                "SELECT id FROM qr_codes WHERE code = ? LIMIT 1",
                (code,),
            ).fetchone()

            cur.execute("""
                INSERT INTO qr_codes (code, points, is_active)
                VALUES (?, ?, 1)
                ON CONFLICT(code) DO UPDATE SET
                    points = excluded.points,
                    is_active = 1
            """, (code, points))

            if old:
                updated += 1
            else:
                added += 1

        conn.commit()

    return {
        "added": added,
        "updated": updated,
        "skipped": skipped,
        "total": len(codes),
    }


def seed_default_qr_codes():
    raw = load_qr_codes_raw_from_folder()
    codes = parse_default_qr_codes(raw)

    if not codes:
        logging.warning("No QR/Data Matrix codes found in %s", QR_CODES_DIR)
        return

    result = upsert_qr_codes_to_db(codes)

    logging.info(
        "QR/Data Matrix codes synced: total=%s added=%s updated=%s skipped=%s",
        result["total"],
        result["added"],
        result["updated"],
        result["skipped"],
    )


def ensure_loyalty_user(
    conn: sqlite3.Connection,
    user_id: int,
    language: str = "ru",
    username: str = "",
    full_name: str = "",
):
    username = normalize_username(username)
    full_name = (full_name or "").strip()

    conn.execute("""
        INSERT INTO loyalty_users (user_id, username, full_name, language, bonus_balance)
        VALUES (?, NULLIF(?, ''), NULLIF(?, ''), ?, 0)
        ON CONFLICT(user_id) DO UPDATE SET
            username = COALESCE(NULLIF(excluded.username, ''), loyalty_users.username),
            full_name = COALESCE(NULLIF(excluded.full_name, ''), loyalty_users.full_name),
            language = excluded.language,
            updated_at = CURRENT_TIMESTAMP
    """, (user_id, username, full_name, language))


def get_user_balance(conn: sqlite3.Connection, user_id: int) -> int:
    row = conn.execute(
        "SELECT bonus_balance FROM loyalty_users WHERE user_id = ?",
        (user_id,),
    ).fetchone()

    return int(row["bonus_balance"]) if row else 0


def find_qr_code(conn: sqlite3.Connection, code: str):
    return conn.execute(
        "SELECT id, code, points, is_active FROM qr_codes WHERE code = ? LIMIT 1",
        (code,),
    ).fetchone()


def is_qr_used(conn: sqlite3.Connection, code: str) -> bool:
    return conn.execute(
        "SELECT 1 FROM qr_code_redemptions WHERE code = ? LIMIT 1",
        (code,),
    ).fetchone() is not None


def redeem_qr_code(
    conn: sqlite3.Connection,
    user_id: int,
    code: str,
    language: str,
    username: str = "",
    full_name: str = "",
):
    ensure_loyalty_user(conn, user_id, language, username=username, full_name=full_name)
    conn.commit()

    qr_row = find_qr_code(conn, code)

    if not qr_row or int(qr_row["is_active"]) != 1:
        return {
            "status": "invalid",
            "points": 0,
            "balance": get_user_balance(conn, user_id),
        }

    if is_qr_used(conn, code):
        return {
            "status": "used",
            "points": 0,
            "balance": get_user_balance(conn, user_id),
        }

    points = int(qr_row["points"])

    try:
        conn.execute("BEGIN IMMEDIATE")

        if conn.execute(
            "SELECT 1 FROM qr_code_redemptions WHERE code = ? LIMIT 1",
            (code,),
        ).fetchone():
            conn.rollback()
            return {
                "status": "used",
                "points": 0,
                "balance": get_user_balance(conn, user_id),
            }

        conn.execute(
            "INSERT INTO qr_code_redemptions (code, user_id, points) VALUES (?, ?, ?)",
            (code, user_id, points),
        )

        conn.execute("""
            UPDATE loyalty_users
            SET bonus_balance = bonus_balance + ?,
                language = ?,
                updated_at = CURRENT_TIMESTAMP
            WHERE user_id = ?
        """, (points, language, user_id))

        conn.execute("""
            INSERT INTO bonus_transactions (user_id, amount, type, code, comment)
            VALUES (?, ?, 'accrual', ?, 'QR bonus accrual')
        """, (user_id, points, code))

        conn.commit()

    except sqlite3.IntegrityError:
        conn.rollback()
        return {
            "status": "used",
            "points": 0,
            "balance": get_user_balance(conn, user_id),
        }

    except Exception:
        conn.rollback()
        raise

    return {
        "status": "ok",
        "points": points,
        "balance": get_user_balance(conn, user_id),
    }


def add_qr_code(code: str, points: int = 1) -> bool:
    with closing(get_db_connection()) as conn:
        cur = conn.cursor()

        cur.execute(
            "INSERT OR IGNORE INTO qr_codes (code, points, is_active) VALUES (?, ?, 1)",
            (code.strip(), int(points)),
        )

        conn.commit()
        return cur.rowcount > 0


def delete_qr_code(code: str) -> bool:
    with closing(get_db_connection()) as conn:
        cur = conn.cursor()
        cur.execute("DELETE FROM qr_codes WHERE code = ?", (code.strip(),))
        conn.commit()
        return cur.rowcount > 0


def list_qr_codes(limit: int = 100):
    with closing(get_db_connection()) as conn:
        return conn.execute(
            "SELECT code, points, is_active, created_at FROM qr_codes ORDER BY id DESC LIMIT ?",
            (limit,),
        ).fetchall()


def ensure_user_exists(
    user_id: int,
    language: str = "ru",
    username: str = "",
    full_name: str = "",
):
    with closing(get_db_connection()) as conn:
        ensure_loyalty_user(conn, user_id, language, username=username, full_name=full_name)
        conn.commit()


def admin_add_bonus(user_id: int, amount: int, language: str = "ru") -> int:
    with closing(get_db_connection()) as conn:
        ensure_loyalty_user(conn, user_id, language)

        conn.execute("""
            UPDATE loyalty_users
            SET bonus_balance = bonus_balance + ?,
                updated_at = CURRENT_TIMESTAMP
            WHERE user_id = ?
        """, (amount, user_id))

        conn.execute("""
            INSERT INTO bonus_transactions (user_id, amount, type, comment)
            VALUES (?, ?, 'manual', 'Admin manual bonus')
        """, (user_id, amount))

        conn.commit()
        return get_user_balance(conn, user_id)


def get_balance_by_user_id(user_id: int, language: str = "ru") -> int:
    with closing(get_db_connection()) as conn:
        ensure_loyalty_user(conn, user_id, language)
        conn.commit()
        return get_user_balance(conn, user_id)


def find_client_by_username(username: str) -> Optional[sqlite3.Row]:
    username = normalize_username(username)

    if not username:
        return None

    with closing(get_db_connection()) as conn:
        return conn.execute(
            """
            SELECT user_id, username, full_name, language, bonus_balance, created_at, updated_at
            FROM loyalty_users
            WHERE lower(username) = ?
            LIMIT 1
            """,
            (username,),
        ).fetchone()


def format_client_name(row: sqlite3.Row) -> str:
    full_name = (row["full_name"] or "").strip()
    if full_name:
        return full_name

    username = (row["username"] or "").strip()
    if username:
        return f"@{username}"

    return str(row["user_id"])


async def process_scanned_qr_code(message: Message, code: str, language: str) -> None:
    user_id = message.from_user.id if message.from_user else 0

    try:
        with closing(get_db_connection()) as conn:
            result = redeem_qr_code(conn, user_id, code, language, username=(message.from_user.username if message.from_user else ""), full_name=make_full_name(message))

        status = result["status"]
        points = int(result.get("points", 0))
        balance = int(result.get("balance", 0))

        if status == "ok":
            title = scan_t(language, "scan_ok_title")
            await message.answer(
                scan_t(language, "scan_ok_text", title=title, code=code, points=points, balance=balance)
            )
            return

        if status == "used":
            title = scan_t(language, "scan_used_title")
            await message.answer(
                scan_t(language, "scan_used_text", title=title, code=code)
            )
            return

        title = scan_t(language, "scan_invalid_title")
        await message.answer(
            scan_t(language, "scan_invalid_text", title=title, code=code)
        )

    except Exception as e:
        logging.exception("process_scanned_qr_code error: %s", e)
        await message.answer(t(language, "processing_error"))


async def import_uploaded_qr_file(message: Message, document: Document):
    user_id = message.from_user.id if message.from_user else 0
    lang = get_user_lang(message)

    if not is_admin(user_id):
        await message.answer(t(lang, "admin_only"))
        return

    file_name = safe_file_name(document.file_name or "")
    lower_name = file_name.lower()

    if not lower_name.endswith((".txt", ".xlsx")):
        await message.answer("❌ Можно загружать только файлы .txt или .xlsx")
        return

    if lower_name.endswith(".xlsx") and load_workbook is None:
        await message.answer(
            "❌ XLSX импорт недоступен: пакет openpyxl не установлен.\n"
            "Добавь <code>openpyxl==3.1.5</code> в requirements.txt или загрузи TXT файл."
        )
        return

    ensure_qr_codes_folder()

    uploads_dir = os.path.join(QR_CODES_DIR, "uploads")
    os.makedirs(uploads_dir, exist_ok=True)

    saved_path = os.path.join(uploads_dir, f"{int(time.time())}_{file_name}")

    try:
        file = await bot.get_file(document.file_id)
        await bot.download_file(file.file_path, destination=saved_path)

        if lower_name.endswith(".txt"):
            with open(saved_path, "r", encoding="utf-8") as f:
                content = f.read()

            codes = parse_qr_txt_content(content)
        else:
            codes = parse_qr_xlsx_file(saved_path)

        if not codes:
            await message.answer(
                "❌ В файле не найдены QR-коды.\n\n"
                "TXT формат:\n"
                "<code>КОД 5 баллов</code>\n\n"
                "XLSX формат:\n"
                "1 колонка — код\n"
                "2 колонка — баллы"
            )
            return

        result = upsert_qr_codes_to_db(codes)

        await message.answer(
            "✅ Импорт завершён\n\n"
            f"Файл: <code>{file_name}</code>\n"
            f"Всего строк: <b>{result['total']}</b>\n"
            f"Добавлено: <b>{result['added']}</b>\n"
            f"Обновлено: <b>{result['updated']}</b>\n"
            f"Пропущено: <b>{result['skipped']}</b>\n\n"
            "База обновлена без перезапуска бота."
        )

    except Exception as e:
        logging.exception("QR file import error: %s", e)
        await message.answer("❌ Ошибка импорта файла.")


@dp.message(CommandStart())
async def start_handler(message: Message):
    lang = get_user_lang(message)
    ensure_user_exists(message.from_user.id if message.from_user else 0, lang, username=(message.from_user.username if message.from_user else ""), full_name=make_full_name(message))

    await message.answer(
        t(lang, "welcome"),
        reply_markup=build_main_keyboard(lang),
    )


@dp.message(Command("help"))
async def help_handler(message: Message):
    await message.answer(t(get_user_lang(message), "help"))


@dp.message(Command("id"))
async def id_handler(message: Message):
    lang = get_user_lang(message)
    user_id = message.from_user.id if message.from_user else 0

    await message.answer(
        t(
            lang,
            "your_id",
            user_id=user_id,
            admin="YES" if is_admin(user_id) else "NO",
        )
    )


@dp.message(Command("debug_url"))
async def debug_url_handler(message: Message):
    lang = get_user_lang(message)
    user_id = message.from_user.id if message.from_user else 0

    if not is_admin(user_id):
        await message.answer(t(lang, "admin_only"))
        return

    await message.answer(t(lang, "debug_url", url=WEBAPP_URL))


@dp.message(Command("addqr"))
async def addqr_handler(message: Message):
    lang = get_user_lang(message)
    user_id = message.from_user.id if message.from_user else 0

    if not is_admin(user_id):
        await message.answer(t(lang, "admin_only"))
        return

    parts = (message.text or "").split(maxsplit=2)

    if len(parts) < 3:
        await message.answer(t(lang, "cmd_usage_addqr"))
        return

    code = parts[1].strip()

    try:
        points = int(parts[2].strip())
    except Exception:
        await message.answer(t(lang, "invalid_number"))
        return

    if points <= 0:
        await message.answer(t(lang, "invalid_number"))
        return

    try:
        ok = add_qr_code(code, points)

        if ok:
            await message.answer(t(lang, "addqr_ok", code=code, points=points))
        else:
            await message.answer(t(lang, "addqr_exists", code=code))

    except Exception as e:
        logging.exception("addqr error: %s", e)
        await message.answer(t(lang, "internal_error"))


@dp.message(Command("delqr"))
async def delqr_handler(message: Message):
    lang = get_user_lang(message)
    user_id = message.from_user.id if message.from_user else 0

    if not is_admin(user_id):
        await message.answer(t(lang, "admin_only"))
        return

    parts = (message.text or "").split(maxsplit=1)

    if len(parts) < 2:
        await message.answer(t(lang, "cmd_usage_delqr"))
        return

    code = parts[1].strip()

    try:
        ok = delete_qr_code(code)

        if ok:
            await message.answer(t(lang, "delqr_ok", code=code))
        else:
            await message.answer(t(lang, "delqr_not_found", code=code))

    except Exception as e:
        logging.exception("delqr error: %s", e)
        await message.answer(t(lang, "internal_error"))


@dp.message(Command("listqr"))
async def listqr_handler(message: Message):
    lang = get_user_lang(message)
    user_id = message.from_user.id if message.from_user else 0

    if not is_admin(user_id):
        await message.answer(t(lang, "admin_only"))
        return

    try:
        rows = list_qr_codes(limit=100)

        if not rows:
            await message.answer(t(lang, "listqr_empty"))
            return

        items = []

        for row in rows:
            items.append(
                f"• <code>{row['code']}</code>\n"
                f"  points: <b>{row['points']}</b> | "
                f"{'active' if int(row['is_active']) == 1 else 'inactive'}"
            )

        await message.answer(
            t(lang, "listqr_header", items="\n\n".join(items))
        )

    except Exception as e:
        logging.exception("listqr error: %s", e)
        await message.answer(t(lang, "internal_error"))


@dp.message(Command("syncqr"))
async def syncqr_handler(message: Message):
    lang = get_user_lang(message)
    user_id = message.from_user.id if message.from_user else 0

    if not is_admin(user_id):
        await message.answer(t(lang, "admin_only"))
        return

    try:
        raw = load_qr_codes_raw_from_folder()
        codes = parse_default_qr_codes(raw)

        if not codes:
            await message.answer("❌ В папке data/qr_codes не найдены QR-коды.")
            return

        result = upsert_qr_codes_to_db(codes)

        await message.answer(
            "✅ QR-коды обновлены из папки <code>data/qr_codes</code>\n\n"
            f"Всего: <b>{result['total']}</b>\n"
            f"Добавлено: <b>{result['added']}</b>\n"
            f"Обновлено: <b>{result['updated']}</b>\n"
            f"Пропущено: <b>{result['skipped']}</b>\n\n"
            "Перезапуск бота не нужен."
        )

    except Exception as e:
        logging.exception("syncqr error: %s", e)
        await message.answer(t(lang, "internal_error"))


@dp.message(Command("importqr"))
async def importqr_handler(message: Message):
    lang = get_user_lang(message)
    user_id = message.from_user.id if message.from_user else 0

    if not is_admin(user_id):
        await message.answer(t(lang, "admin_only"))
        return

    await message.answer(
        "📥 Импорт QR-кодов\n\n"
        "Можно отправить боту файл <b>.txt</b> или <b>.xlsx</b>.\n\n"
        "TXT формат:\n"
        "<code>047000280627027ABC 5 баллов</code>\n\n"
        "XLSX формат:\n"
        "1 колонка — QR/Data Matrix код\n"
        "2 колонка — количество баллов\n\n"
        "После загрузки файла база обновится автоматически без перезапуска бота."
    )


@dp.message(Command("bonus"))
async def bonus_handler(message: Message):
    lang = get_user_lang(message)
    user_id = message.from_user.id if message.from_user else 0

    if not is_admin(user_id):
        await message.answer(t(lang, "admin_only"))
        return

    parts = (message.text or "").split(maxsplit=2)

    if len(parts) < 3:
        await message.answer(t(lang, "cmd_usage_bonus"))
        return

    try:
        target_user_id = int(parts[1].strip())
        amount = int(parts[2].strip())
    except Exception:
        await message.answer(t(lang, "invalid_number"))
        return

    if amount == 0:
        await message.answer(t(lang, "invalid_number"))
        return

    try:
        balance = admin_add_bonus(target_user_id, amount, lang)

        await message.answer(
            t(
                lang,
                "bonus_ok",
                user_id=target_user_id,
                amount=amount,
                balance=balance,
            )
        )

    except Exception as e:
        logging.exception("bonus error: %s", e)
        await message.answer(t(lang, "internal_error"))


@dp.message(Command("balance"))
async def balance_handler(message: Message):
    lang = get_user_lang(message)
    user_id = message.from_user.id if message.from_user else 0

    if not is_admin(user_id):
        await message.answer(t(lang, "admin_only"))
        return

    parts = (message.text or "").split(maxsplit=1)

    if len(parts) < 2:
        await message.answer(t(lang, "cmd_usage_balance"))
        return

    try:
        target_user_id = int(parts[1].strip())
    except Exception:
        await message.answer(t(lang, "invalid_number"))
        return

    try:
        balance = get_balance_by_user_id(target_user_id, lang)

        await message.answer(
            t(
                lang,
                "balance_text",
                user_id=target_user_id,
                balance=balance,
            )
        )

    except Exception as e:
        logging.exception("balance error: %s", e)
        await message.answer(t(lang, "internal_error"))



@dp.message(Command("check"))
async def check_client_handler(message: Message):
    lang = get_user_lang(message)
    admin_user_id = message.from_user.id if message.from_user else 0
    update_user_profile(message, lang)

    if not is_admin(admin_user_id):
        await message.answer(t(lang, "admin_only"))
        return

    parts = (message.text or "").split(maxsplit=1)

    if len(parts) < 2:
        await message.answer(t(lang, "cmd_usage_check"))
        return

    username = normalize_username(parts[1])

    if not username:
        await message.answer(t(lang, "cmd_usage_check"))
        return

    try:
        row = find_client_by_username(username)

        if not row:
            await message.answer(t(lang, "client_not_found", username=username))
            return

        balance = int(row["bonus_balance"] or 0)
        status = get_client_status(balance)

        await message.answer(
            t(
                lang,
                "client_info",
                full_name=format_client_name(row),
                username=row["username"] or username,
                user_id=row["user_id"],
                balance=balance,
                status=status,
            )
        )

    except Exception as e:
        logging.exception("check client error: %s", e)
        await message.answer(t(lang, "internal_error"))


@dp.message(Command("status"))
async def status_client_handler(message: Message):
    lang = get_user_lang(message)
    admin_user_id = message.from_user.id if message.from_user else 0
    update_user_profile(message, lang)

    if not is_admin(admin_user_id):
        await message.answer(t(lang, "admin_only"))
        return

    parts = (message.text or "").split(maxsplit=1)

    if len(parts) < 2:
        await message.answer(t(lang, "cmd_usage_check"))
        return

    username = normalize_username(parts[1])

    if not username:
        await message.answer(t(lang, "cmd_usage_check"))
        return

    try:
        row = find_client_by_username(username)

        if not row:
            await message.answer(t(lang, "client_not_found", username=username))
            return

        balance = int(row["bonus_balance"] or 0)
        status = get_client_status(balance)

        await message.answer(
            t(
                lang,
                "client_status",
                full_name=format_client_name(row),
                username=row["username"] or username,
                balance=balance,
                status=status,
            )
        )

    except Exception as e:
        logging.exception("status client error: %s", e)
        await message.answer(t(lang, "internal_error"))


@dp.message(F.document)
async def document_handler(message: Message):
    if not message.document:
        return

    await import_uploaded_qr_file(message, message.document)


@dp.message(F.web_app_data)
async def handle_web_app_data(message: Message):
    raw_data = message.web_app_data.data if message.web_app_data else ""

    logging.info("WEB_APP_DATA RAW: %s", raw_data)

    try:
        data = json.loads(raw_data)
    except Exception as e:
        logging.exception("WEB_APP_DATA JSON parse error: %s", e)
        await message.answer(t(get_user_lang(message), "bad_data"))
        return

    action = str(data.get("action", "")).strip()
    code = str(data.get("code", "")).strip()
    language = normalize_lang(str(data.get("language", "ru")).strip().lower() or "ru")

    if action != "scan_qr":
        await message.answer(t(language, "unknown_action"))
        return

    if not code:
        await message.answer(t(language, "empty_code"))
        return

    user_id = message.from_user.id if message.from_user else 0
    update_user_profile(message, language)

    if is_duplicate_scan(user_id, code):
        logging.info("Duplicate scan ignored: user_id=%s code=%s", user_id, code)
        return

    await process_scanned_qr_code(message, code, language)


async def set_menu_button():
    try:
        await bot.set_chat_menu_button(
            menu_button=MenuButtonWebApp(
                text=TEXTS["ru"]["menu_button"],
                web_app=WebAppInfo(url=WEBAPP_URL),
            )
        )
        logging.info("Menu button WebApp set successfully")
    except Exception as e:
        logging.warning("Menu button setup failed: %s", e)


async def on_startup():
    logging.info("Skipping delete_webhook to avoid Telegram timeout")

    try:
        init_qr_bonus_tables()
        seed_default_qr_codes()
    except Exception as e:
        logging.exception("Database init error: %s", e)

    try:
        await set_menu_button()
    except Exception as e:
        logging.warning("Menu button setup failed: %s", e)

    logging.info("Bot started successfully")


async def on_shutdown():
    try:
        await bot.session.close()
    except Exception:
        pass

    logging.info("Bot stopped")


async def main():
    await on_startup()

    try:
        await dp.start_polling(
            bot,
            polling_timeout=60,
            allowed_updates=dp.resolve_used_update_types(),
            handle_signals=True,
            close_bot_session=False,
        )
    finally:
        await on_shutdown()


if __name__ == "__main__":
    try:
        asyncio.run(main())
    except (KeyboardInterrupt, SystemExit):
        logging.info("Application interrupted")